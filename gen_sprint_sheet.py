"""Generate the HR Recruited sprint workbook in the Rezolv24 "04" style.

Backend-only project, so tracks are split by EPIC (RANK, PIPE) rather than FE/BE.
Follows SPRINT_SHEET_SPEC.md exactly.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

HEADERS = ['Epic', 'User Story #', 'Task #', 'Epic Name', 'User Story', 'Activity',
           'Deliverable', 'Dependent Task', 'Ext. Dependency', 'Int. Dependency', 'ETA']

NAVY, PURPLE, LAVEND, WHITE, INK = '1A1A3E', '2D2D6B', 'EBEBF9', 'FFFFFF', '1A1A3E'
HDR_FILL = EPIC_FILL = PatternFill('solid', fgColor=NAVY)
US_FILL = PatternFill('solid', fgColor=PURPLE)
LAVEND_FILL = PatternFill('solid', fgColor=LAVEND)
WHITE_FILL = PatternFill('solid', fgColor=WHITE)
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
US_FONT = Font(bold=True, color='FFFFFF', size=10)
TASK_FONT = Font(bold=False, color=INK, size=9)
ALIGN = Alignment(wrap_text=True, vertical='center', horizontal='left')
WIDTHS = {'A': 8, 'B': 14, 'C': 13, 'D': 32, 'E': 36, 'F': 38, 'G': 72, 'H': 28, 'I': 26, 'J': 22, 'K': 7}


def _style(ws, row, fill, font, height):
    for col in range(1, 12):
        c = ws.cell(row=row, column=col)
        c.fill, c.font, c.alignment = fill, font, ALIGN
    ws.row_dimensions[row].height = height


def build_sheet(ws, rows):
    ws.append(HEADERS)
    _style(ws, 1, HDR_FILL, HDR_FONT, 36)
    stripe = 0
    for r in rows:
        if r[0] == 'epic':
            ws.append([r[1], None, None, r[2]])
            _style(ws, ws.max_row, EPIC_FILL, HDR_FONT, 26)
        elif r[0] == 'us':
            ws.append([None, r[1], None, None, r[2]])
            _style(ws, ws.max_row, US_FILL, US_FONT, 22)
            stripe = 0
        else:  # task: supports optional int-dependency at r[6]
            _, tid, title, deliverable, dep, eta = r[:6]
            intdep = r[6] if len(r) > 6 else None
            ws.append([None, None, tid, None, None, title, deliverable, dep, None, intdep, eta])
            fill = LAVEND_FILL if stripe % 2 == 0 else WHITE_FILL
            _style(ws, ws.max_row, fill, TASK_FONT, 60)
            stripe += 1
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w


RANK = [
    ('epic', 'EP-RANK', 'Ranking Agent (Agent 5) — Backend'),

    ('us', 'US-RANK-BE-01',
     'As a recruiter I get one combined score per candidate that blends CV match and interview performance.'),
    ('task', 'EP-RANK-BE-001', 'Composite scoring core',
     'New module ranking_agent.py. Implement pure fn compute_composite(fit_percent, interview_score) returning '
     '{composite_score, score_breakdown {cv_match, interview, weights}}. Weights read from env RANK_CV_WEIGHT '
     '(default 0.4) and RANK_INTERVIEW_WEIGHT (default 0.6), normalised to sum 1.0. Clamp result 0-100. When '
     'interview_score is None (candidate never interviewed) score CV-only so half-scored candidates cannot '
     'outrank interviewed ones. No DB, no side effects.', '—', '2h'),
    ('task', 'EP-RANK-BE-002', 'Recommendation mapping',
     'Pure fn recommend(composite_score, red_flags) -> strong_yes (>=80) / yes (>=65) / maybe (>=50) / no (else). '
     'When red_flags is non-empty, downgrade: return "review" if score>=70 else "no". Guard clause on empty input.',
     'EP-RANK-BE-001', '1h'),
    ('task', 'EP-RANK-BE-003', 'Module self-test block',
     'Add an if __name__ == "__main__" block exercising compute_composite happy path, the None-interview CV-only '
     'path, and recommend across strong_yes / review (with flag) / no — per repo convention that every module '
     'self-tests. No DB required to run.', 'EP-RANK-BE-002', '0.5h'),

    ('us', 'US-RANK-BE-02',
     'As a recruiter the system persists a unified ranked shortlist I can pull per job.'),
    ('task', 'EP-RANK-BE-004', 'Persist per-candidate ranking',
     'Add collection helpers (_get_db/_candidates/_screened/_insights/_ranked) matching the lazy-singleton pattern '
     'in shortlisting_agent.py. Implement rank_candidate(cv_id): read fit_percent from candidates_info, '
     'interview_score from screened_candidates, red_flags/key_strengths from interview_insights; compute composite '
     '+ recommendation; upsert a ranked_candidates doc {_id, jd_id, name/email/phone, composite_score, '
     'score_breakdown, recommendation, red_flags, key_strengths, interview_status, ranked_at}; mirror '
     'composite_score+recommendation onto candidates_info. Guard-clause return None on unknown cv_id. Idempotent.',
     'EP-RANK-BE-001', '3h'),
    ('task', 'EP-RANK-BE-005', 'Ranked shortlist query',
     'Implement rank_for_jd(jd_id, top_n=None) returning ranked_candidates for the job sorted by composite_score '
     'descending, each annotated with a 1-based rank field; top_n caps the list, omitting returns the full list.',
     'EP-RANK-BE-004', '2h'),
    ('task', 'EP-RANK-BE-006', 'Ranked-candidates API endpoint',
     'Add GET /ranked-candidates to app.py (query params jd_id, optional top_n, skip/limit pagination consistent '
     'with the existing /shortlisted-candidates endpoint). Add a RankedListResponse Pydantic model to schemas.py. '
     'Return ranked records highest-composite-first.', 'EP-RANK-BE-005', '3h'),

    ('us', 'US-RANK-BE-03',
     'As a recruiter I can re-rank a job on demand after changing the scoring weights.'),
    ('task', 'EP-RANK-BE-007', 'On-demand re-rank endpoint',
     'Add POST /jobs/{jd_id}/rerank to app.py that recomputes every ranked_candidates record for the JD using the '
     'current env weights (PRD "re-ranking flow"); optionally accept {cv_weight, interview_weight} in the request '
     'body to override for that run. Returns the refreshed ranked shortlist.', 'EP-RANK-BE-004', '2h'),
]

PIPE = [
    ('epic', 'EP-PIPE', 'Cohesive LangGraph Pipeline (Agents 2-5) — Backend'),

    ('us', 'US-PIPE-BE-01',
     'As the system, each candidate flows through one connected, checkpointed pipeline graph instead of separate scheduler steps.'),
    ('task', 'EP-PIPE-BE-001', 'Define shared pipeline state',
     'New module pipeline.py. Define PipelineState TypedDict: cv_id, jd_id, profile fields (name/email/phone/'
     'experience/education/raw_cv_text), fit_percent, call_id/transcript/summary/interview_score/call_category, '
     'booking_token/selected_slot/meet_link, composite_score/recommendation, stage/status. Must be JSON-'
     'serialisable so the checkpointer can persist it.', '—', '3h'),
    ('task', 'EP-PIPE-BE-002', 'Checkpointer + graph factory',
     'Add create_pipeline_agent() that compiles an (initially empty) StateGraph over PipelineState with a SQLite '
     'checkpointer, falling back to InMemorySaver on failure — reuse the pattern from job_post.py:_build_'
     'checkpointer (separate sqlite file, e.g. pipeline_checkpoints.sqlite).', 'EP-PIPE-BE-001', '2h'),

    ('us', 'US-PIPE-BE-02',
     'As the system, a new applicant is parsed and scored against their JD, then routed by CV fit.'),
    ('task', 'EP-PIPE-BE-003', 'Parse CV node',
     'Add parse_cv node wrapping parser_agent.process_cv(pdf_path, jd_id); populate the profile fields of '
     'PipelineState from the returned candidate dict. Node runs in the first invoke before any interrupt, so the '
     'temp PDF path is valid. Set entry point to parse_cv.', 'EP-PIPE-BE-002', '3h'),
    ('task', 'EP-PIPE-BE-004', 'Match node',
     'Add match node wrapping shortlisting_agent._score_candidate_against_jd(candidate, jd_text) (load jd_text '
     'from job_descriptions by jd_id); write fit_percent into both PipelineState and candidates_info to keep the '
     'existing dashboards in sync.', 'EP-PIPE-BE-003', '2h'),
    ('task', 'EP-PIPE-BE-005', 'CV-threshold router + reject node',
     'Add conditional edge route_after_match: fit_percent >= VOICE_CALL_THRESHOLD (env) routes to start_interview, '
     'otherwise to a terminal reject node that records status="rejected_cv" and ends. Read the threshold from env, '
     'no hardcode.', 'EP-PIPE-BE-004', '2h'),

    ('us', 'US-PIPE-BE-03',
     'As the system, a qualified candidate is voice-interviewed and the call result resumes the pipeline.'),
    ('task', 'EP-PIPE-BE-006', 'Interview node + interrupt',
     'Add start_interview node: call _initiate_vapi_call(phone_e164), insert the screened_candidates row '
     '(status="calling") exactly as call_top_candidates does today, then interrupt() to suspend the graph until '
     'the Vapi webhook arrives. Highest-risk ticket (async interrupt/resume) — assign to strongest dev.',
     'EP-PIPE-BE-005', '4h'),
    ('task', 'EP-PIPE-BE-007', 'Webhook resume wiring',
     'Modify POST /voice/webhook in app.py to resolve cv_id via screened_candidates.call_id and resume the graph '
     'thread with Command(resume={transcript, summary, end_reason, duration}) instead of calling record_call_'
     'result directly. Guard unknown/duplicate call_id (no crash). No regression to the status="calling" insert.',
     'EP-PIPE-BE-006', '4h'),
    ('task', 'EP-PIPE-BE-008', 'Interview scoring node',
     'Add score_interview node (resume target): categorize_call, then _score_interview(transcript, summary, '
     'job_desc) and extract_interview_insights; persist interview_score, call_category and insights with parity to '
     'the current record_call_result behaviour.', 'EP-PIPE-BE-007', '3h'),
    ('task', 'EP-PIPE-BE-009', 'Interview-threshold router',
     'Add route_after_interview: completed AND interview_score >= CALENDAR_THRESHOLD -> book; completed AND below '
     '-> rank; not-completed -> reject (keep the call_logs retry path compatible). All three branches reachable.',
     'EP-PIPE-BE-008', '2h'),

    ('us', 'US-PIPE-BE-04',
     'As the system, a strong interviewee is sent a slot picker and their selection resumes the pipeline.'),
    ('task', 'EP-PIPE-BE-010', 'Booking node + interrupt',
     'Add book node: call create_slot_picker_booking(candidate_doc, score) to email the picker link, then '
     'interrupt() to suspend until the candidate selects a slot.', 'EP-PIPE-BE-009', '4h'),
    ('task', 'EP-PIPE-BE-011', 'Slot-select resume + event node',
     'Modify POST /api/booking/{token}/select in app.py to resolve cv_id via the pending_bookings token and resume '
     'the thread with Command(resume={selected_slot, meet_link}); add create_event node that records meet_link/'
     'selected_slot into state (preserve existing select_slot calendar-event behaviour).', 'EP-PIPE-BE-010', '4h'),

    ('us', 'US-PIPE-BE-05',
     'As a recruiter, every candidate who completes the journey ends with a composite ranked record.'),
    ('task', 'EP-PIPE-BE-012', 'Ranking terminal node',
     'Add rank terminal node calling ranking_agent.rank_candidate(cv_id) then END. Both the booking path and the '
     'below-booking-threshold completed path converge here. Cross-epic dependency: requires EP-RANK-BE-004.',
     'EP-PIPE-BE-011', '2h', 'EP-RANK-BE-004'),

    ('us', 'US-PIPE-BE-06',
     'As the system, the scheduler launches one pipeline run per applicant instead of running ingest/shortlist/call sequentially.'),
    ('task', 'EP-PIPE-BE-013', 'Scheduler launches pipeline',
     'Refactor _shortlist_loop (app.py:34) so it only detects new applicants (new Sheet rows / Drive file_ids, '
     'dedup via processed_applications) and launches or resumes one graph per candidate; stop calling shortlist_'
     'all_jobs / call_top_candidates directly — the graph now owns scoring and calling.', 'EP-PIPE-BE-012', '4h'),
    ('task', 'EP-PIPE-BE-014', 'End-to-end validation harness',
     'Extend the test_booking_flow.py-style standalone harness to drive one fixture candidate from parse to a '
     'ranked_candidates record, simulating the Vapi webhook and slot-pick resumes; assert existing /candidates, '
     '/shortlisted-candidates, /screened_candidates endpoints still read the same collections.', 'EP-PIPE-BE-013',
     '4h'),
]


def _eta_total(rows):
    tot = 0.0
    for r in rows:
        if r[0] == 'task':
            tot += float(r[5].rstrip('h'))
    return tot


wb = openpyxl.Workbook()
s1 = wb.active
s1.title = 'Backend — RANK'
build_sheet(s1, RANK)
s2 = wb.create_sheet('Backend — PIPE')
build_sheet(s2, PIPE)
wb.save('HRRecruited_Sprint_RANK_PIPE.xlsx')

print(f"RANK total: {_eta_total(RANK)}h ({sum(1 for r in RANK if r[0]=='task')} tasks)")
print(f"PIPE total: {_eta_total(PIPE)}h ({sum(1 for r in PIPE if r[0]=='task')} tasks)")
print("saved HRRecruited_Sprint_RANK_PIPE.xlsx")
